"""Voter-facing routes — round-aware."""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, Request, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, ValidationError, conint
from sqlalchemy.orm import Session, joinedload
import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ballot.auth import require_voter
from ballot.database import get_db
from ballot.models import (
    Nomination, NominationType,
    Nominee, Vote, Ranking, Voter,
    Round, RoundParticipation, RoundType,
    ContestNomination, NominationTemplate,
)

router = APIRouter(dependencies=[Depends(require_voter)])
templates = Jinja2Templates(directory="ballot/templates")
SAFE_FILENAME_RE = re.compile(r"[^0-9A-Za-z_-]")
MAX_EXPORT_NOMINATIONS = 100
MAX_EXPORT_ROWS = 500
MAX_EXPORT_COLS = 10


class DraftModel(BaseModel):
    picks: dict[str, list[conint(ge=1)]] = {}
    runnerups: dict[str, conint(ge=1)] = {}
    ranks: dict[str, dict[str, conint(ge=1)]] = {}


class ExportRowModel(BaseModel):
    cols: list[str]
    urls: list[str | None] = []


class ExportNominationModel(BaseModel):
    name: str
    header: list[str]
    rows: list[ExportRowModel]


class BallotExportModel(BaseModel):
    nominations: list[ExportNominationModel] = []


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _content_disposition(filename: str) -> str:
    """
    Generate Content-Disposition header value for file download.

    Creates a properly formatted Content-Disposition header that works
    with both ASCII and UTF-8 filenames for browser downloads.

    :param filename: Original filename to be downloaded
    :returns: Formatted Content-Disposition header value
    """
    ascii_name = filename.encode("ascii", "ignore").decode()
    encoded_name = quote(filename, safe="")
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"


def _safe_filename_label(raw: str) -> str:
    """
    Create a safe filename label from raw input.

    Strips unsafe characters and limits length to prevent filesystem issues.

    :param raw: Raw string to convert to safe filename
    :returns: Safe filename label
    """
    safe = SAFE_FILENAME_RE.sub("_", (raw or "").strip())[:60]
    return safe or "ballot"


def _parse_draft_payload(body: Any) -> dict[str, Any]:
    """
    Parse and validate draft payload from request body.

    Validates the incoming JSON payload against the DraftModel schema
    and returns the validated data as a dictionary.

    :param body: Raw request body data
    :returns: Validated draft data as dictionary
    :raises HTTPException: If payload validation fails
    """
    try:
        return DraftModel.parse_obj(body).dict()
    except ValidationError:
        raise HTTPException(status_code=400, detail="Invalid draft payload")


def _parse_export_payload(body: Any) -> list[dict[str, Any]]:
    """
    Parse and validate export payload from request body.

    Validates the incoming JSON payload against the BallotExportModel schema,
    enforces limits on nominations, rows, and columns, and returns the 
    validated data as a list of dictionaries.

    :param body: Raw request body data
    :returns: Validated export data as list of dictionaries
    :raises HTTPException: If payload validation fails or limits exceeded
    """
    try:
        payload = BallotExportModel.parse_obj(body)
    except ValidationError:
        raise HTTPException(status_code=400, detail="Invalid export payload")

    if len(payload.nominations) > MAX_EXPORT_NOMINATIONS:
        raise HTTPException(status_code=400, detail="Too many nominations in export payload")

    nominations_data: list[dict[str, Any]] = []
    for nom in payload.nominations:
        if len(nom.rows) > MAX_EXPORT_ROWS:
            raise HTTPException(status_code=400, detail="Too many rows in export payload")
        if len(nom.header) > MAX_EXPORT_COLS:
            raise HTTPException(status_code=400, detail="Too many columns in export payload")

        rows: list[dict[str, list[str | None] | list[str]]] = []
        for row in nom.rows:
            if len(row.cols) > MAX_EXPORT_COLS:
                raise HTTPException(status_code=400, detail="Too many columns in export payload")
            rows.append({"cols": row.cols, "urls": row.urls})

        nominations_data.append({
            "name": nom.name,
            "header": nom.header,
            "rows": rows,
        })
    return nominations_data


def _auto_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Auto-adjust column widths in Excel worksheet.

    Calculates optimal column width based on cell content length,
    with a maximum width constraint to prevent excessively wide columns.

    :param ws: OpenPyXL worksheet object
    """
    for col in ws.columns:
        first = next((c for c in col if not isinstance(c, MergedCell)), None)
        if first is None:
            continue
        max_len = max((len(str(c.value or "")) for c in col if not isinstance(c, MergedCell)), default=10)
        ws.column_dimensions[first.column_letter].width = min(max_len + 4, 80)


def _sheet_title(name: str) -> str:
    """
    Sanitize sheet title for Excel compatibility.

    Removes characters that are invalid in Excel sheet names and
    truncates to the maximum allowed length (31 characters).

    :param name: Original sheet name
    :returns: Sanitized sheet title
    """
    for ch in r'\/:*?[]':
        name = name.replace(ch, '')
    return name[:31]


def _write_cell_with_link(
        ws:openpyxl.worksheet.worksheet.Worksheet, 
        row: int, 
        col: int, 
        label: str, 
        url: Optional[str], 
        link_font: Font
    ):
    """
    Write a cell with optional hyperlink and styling.

    Creates a cell with the specified label, and if a URL is provided,
    makes it a clickable hyperlink with styled formatting.

    :param ws: Worksheet object
    :param row: row of the table
    :param col: Column number (1-indexed)
    :param label: Cell text content
    :param url: Optional URL for hyperlink
    :param link_font: Font styling for hyperlinks
    :returns: Configured cell object
    """
    cell = ws.cell(row=row, column=col, value=label)
    if url:
        cell.hyperlink = url
        cell.font = link_font
    return cell


def _build_xlsx_per_nomination(nominations_data: list[dict]) -> io.BytesIO:
    """
    Build an Excel workbook with ballot data organized by nomination.

    Creates a multi-sheet Excel workbook where each nomination becomes a separate sheet.
    Handles proper formatting, hyperlink styling, and auto-width adjustment for readability.

    :param nominations_data: List of nomination data dictionaries with name, header, and rows
    :returns: BytesIO buffer containing the generated Excel workbook
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="4F46E5")
    header_align = Alignment(horizontal="center")
    link_font    = Font(color="1155CC", underline="single")

    seen_titles: dict[str, int] = {}
    for nom in nominations_data:
        raw_title = _sheet_title(nom["name"])
        if raw_title in seen_titles:
            seen_titles[raw_title] += 1
            title = _sheet_title(raw_title)[:28] + f" {seen_titles[raw_title]}"
        else:
            seen_titles[raw_title] = 1
            title = raw_title

        ws = wb.create_sheet(title=title)
        ws.append([nom["name"]])
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=max(len(nom["header"]), 1))
        ws.cell(1, 1).font = Font(bold=True, size=13)
        ws.cell(1, 1).alignment = Alignment(horizontal="left")
        ws.append([])
        ws.append(nom["header"])
        for i in range(1, len(nom["header"]) + 1):
            cell = ws.cell(3, i)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        data_row = 4
        for row in nom["rows"]:
            if isinstance(row, dict):
                cols = row.get("cols", [])
                urls = row.get("urls", [None] * len(cols))
            else:
                cols = row
                urls = [None] * len(cols)
            for col_idx, (val, url) in enumerate(zip(cols, urls), start=1):
                _write_cell_with_link(ws, data_row, col_idx, val, url or None, link_font)
            data_row += 1
        _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _get_or_create_participation(
    db: Session, round_id: int, voter_id: int
) -> RoundParticipation:
    """
    Get or create a participation record for a voter in a round.
    
    :param db: Database session
    :param round_id: ID of the round
    :param voter_id: ID of the voter
    :returns: RoundParticipation object
    """
    p = db.query(RoundParticipation).filter_by(
        round_id=round_id, voter_id=voter_id
    ).first()
    if not p:
        p = RoundParticipation(round_id=round_id, voter_id=voter_id)
        db.add(p)
        db.flush()
    return p


def _nominations_for_round(db: Session, round_id: int) -> list[Nomination]:
    """
    Get nominations for a round.
    
    :param db: Database session
    :param round_id: ID of the round
    :returns: List of Nomination objects
    """
    return (
        db.query(Nomination)
        .options(
            joinedload(Nomination.nominees).joinedload(Nominee.film),
            joinedload(Nomination.nominees).joinedload(Nominee.person),
        )
        .filter(Nomination.round_id == round_id)
        .order_by(Nomination.sort_order, Nomination.id)
        .all()
    )


def _check_cross_nomination_conflict(
    db: Session, 
    voter_id: int, 
    selected_nids: set[int], 
    rnd: Round) -> None:
    """
    Check for cross-nomination conflicts.
    
    :param db: Database session
    :param voter_id: ID of the voter
    :param selected_nids: Set of nominee IDs
    :param rnd: Round object
    """
    if not selected_nids:
        return

    nominees = (
        db.query(Nominee)
        .options(
            joinedload(Nominee.nomination).joinedload(Nomination.contest_nomination).joinedload(ContestNomination.template),
            joinedload(Nominee.persons),
        )
        .filter(Nominee.id.in_(selected_nids))
        .all()
    )

    # (person_id, acting_group) -> set(nomination_id)
    person_group_nominations: dict[tuple[int, str], set[int]] = {}
    # person_id -> name  (для сообщения об ошибке)
    person_names: dict[int, str] = {}

    for nominee in nominees:
        nomination = nominee.nomination
        if not nomination or nomination.round_id != rnd.id:
            continue
        template = nomination.contest_nomination.template if nomination.contest_nomination else None
        acting_group = nomination.acting_group or (template.acting_group if template else None)
        if not acting_group:
            continue

        person_ids = set()
        if nominee.person_id:
            person_ids.add(nominee.person_id)
            if nominee.person:
                person_names[nominee.person_id] = nominee.person.name
        if nominee.persons:
            for np in nominee.persons:
                person_ids.add(np.person_id)
                person_names[np.person_id] = np.person.name

        for pid in person_ids:
            key = (pid, acting_group)
            person_group_nominations.setdefault(key, set()).add(nomination.id)

    conflicts: list[str] = []
    for (pid, acting_group), nomination_ids in person_group_nominations.items():
        if len(nomination_ids) > 1:
            person_name = person_names.get(pid, f"person#{pid}")
            conflicts.append(
                f"«{person_name}» ({acting_group})"
            )

    if conflicts:
        names = ", ".join(conflicts)
        raise HTTPException(
            status_code=400,
            detail=f"Нельзя голосовать за одного актёра в двух номинациях одной группы: {names}",
        )


def _find_active_round_for_year(
    db: Session, year: int, round_type: RoundType | None = None
) -> Round | None:
    """
    Find the active round for the year.
    
    :param db: Database session
    :param year: Year
    :param round_type: Optional round type
    :returns: Round object or None
    """
    q = db.query(Round).filter(
        Round.year == year,
        Round.is_active == True,  # noqa: E712
    )
    if round_type is not None:
        q = q.filter(Round.round_type == round_type)
    rounds = q.all()
    if not rounds:
        return None
    if round_type is not None:
        return rounds[0]
    # prefer FINAL
    for rnd in rounds:
        if rnd.round_type == RoundType.FINAL:
            return rnd
    return rounds[0]


def _find_latest_active_round(db: Session) -> Round | None:
    """
    Find the latest active round.
    
    :param db: Database session
    :returns: Round object or None
    """
    active = (
        db.query(Round)
        .filter(Round.is_active == True)  # noqa: E712
        .all()
    )
    if not active:
        return None
    # Sort: rounds with deadline first (later deadline = higher priority),
    # then no-deadline rounds, prefer FINAL over LONGLIST within same deadline.
    def sort_key(r: Round):
        # None deadline → treat as epoch 0 (lowest priority)
        dl = r.deadline.timestamp() if r.deadline else 0
        is_final = 1 if r.round_type == RoundType.FINAL else 0
        return (dl, is_final, r.year)

    return max(active, key=sort_key)


def _render_vote_page(request, db, rnd, voter):
    """
    Render the vote page.
    
    :param request: FastAPI request object
    :param db: Database session
    :param rnd: Round object
    :param voter: Voter object
    :returns: TemplateResponse object
    """
    nominations = _nominations_for_round(db, rnd.id)
    participation = _get_or_create_participation(db, rnd.id, voter.id)
    db.commit()
    draft = participation.draft or {}
    draft_restored = bool(draft)
    return templates.TemplateResponse(request, "vote.html", {
        "voter": voter,
        "round": rnd,
        "nominations": nominations,
        "draft": draft,
        "draft_restored": draft_restored,
    })


def _deadline_passed(rnd: Round) -> bool:
    """
    Check if the round has a deadline and it has passed.
    
    :param rnd: Round object
    :returns: True if the round has a deadline and it has passed
    """
    if not rnd.deadline:
        return False
    dl = rnd.deadline
    if dl.tzinfo is None:
        # Interpret naive deadline as local server time
        local_tz = datetime.now().astimezone().tzinfo
        dl = dl.replace(tzinfo=local_tz)
    dl_utc = dl.astimezone(timezone.utc)
    return datetime.now(timezone.utc) > dl_utc


def _check_round_open(request, rnd) -> HTMLResponse | None:
    """
    Check if the round is open.
    
    :param request: FastAPI request object
    :param rnd: Round object
    :returns: TemplateResponse object or None
    """
    if not rnd or not rnd.is_active:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": "Этот раунд не активен."},
            status_code=403,
        )
    if _deadline_passed(rnd):
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "round": rnd, "deadline": rnd.deadline, "message": "Дедлайн голосования прошёл."},
            status_code=403,
        )
    return None


# ---------------------------------------------------------------------------
# /vote  — redirects to the most relevant active round
# ---------------------------------------------------------------------------

@router.get("/vote", response_class=HTMLResponse)
def vote_redirect(request: Request, db: Session = Depends(get_db)):
    """
    Redirect to the most relevant active round.
    
    :param request: FastAPI request object
    :param db: Database session
    :returns: RedirectResponse object
    """
    active = _find_latest_active_round(db)
    if not active:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": "Активных раундов нет."},
            status_code=403,
        )
    slug = "final" if active.round_type == RoundType.FINAL else "longlist"
    return RedirectResponse(url=f"/{active.year}/vote/{slug}", status_code=302)


# /{year}/vote  — prefers FINAL, redirects to typed URL
@router.get("/{year}/vote", response_class=HTMLResponse)
def vote_page_year(year: int, request: Request, db: Session = Depends(get_db)):
    """
    Display the vote page for a specific year.
    
    :param year: Year
    :param request: FastAPI request object
    :param db: Database session
    :returns: TemplateResponse object
    """
    rnd = _find_active_round_for_year(db, year)
    if not rnd:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": f"Активных раундов для {year} года нет."},
            status_code=403,
        )
    slug = "final" if rnd.round_type == RoundType.FINAL else "longlist"
    return RedirectResponse(url=f"/{year}/vote/{slug}", status_code=302)


# /{year}/vote/longlist
@router.get("/{year}/vote/longlist", response_class=HTMLResponse)
def vote_longlist(year: int, request: Request, db: Session = Depends(get_db)):
    """
    Displays the voting page for a specific year's LONGLIST round.

    :param year: The year of the active longlist round to display.
    :param request: FastAPI request object providing context for the template response.
    :param db: The SQLAlchemy database session object used to query rounds.
    :return: Redirects to the appropriate voting page based on active round information.
    :rtype: _TemplateResponse
    """
    voter: Voter = request.state.voter
    rnd = _find_active_round_for_year(db, year, RoundType.LONGLIST)
    if not rnd:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": f"Активного лонг-листа для {year} года нет."},
            status_code=403,
        )
    err = _check_round_open(request, rnd)
    if err:
        return err
    return _render_vote_page(request, db, rnd, voter)


@router.post("/{year}/vote/longlist")
async def submit_vote_longlist(year: int, request: Request, db: Session = Depends(get_db)):
    """
    Process submission of votes for a specific year's LONGLIST round.
    
    Validates the active longlist round for the given year, checks if voting is open,
    and processes the submitted votes through the _do_submit function.
    
    :param year: The year of the longlist round for which votes are being submitted.
    :param request: FastAPI request object containing the vote data.
    :param db: The SQLAlchemy database session object.
    :return: Response from _do_submit function processing the vote submission.
    """
    voter: Voter = request.state.voter
    rnd = _find_active_round_for_year(db, year, RoundType.LONGLIST)
    if not rnd:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": f"Активного лонг-листа для {year} года нет."},
            status_code=403,
        )
    err = _check_round_open(request, rnd)
    if err:
        return err
    return await _do_submit(request, db, rnd, voter)


# /{year}/vote/final
@router.get("/{year}/vote/final", response_class=HTMLResponse)
def vote_final(year: int, request: Request, db: Session = Depends(get_db)):
    """
    Displays the voting page for a specific year's FINAL round.

    :param year: The year of the active final round to display.
    :param request: FastAPI request object providing context for the template response.
    :param db: The SQLAlchemy database session object used to query rounds.
    :return: Renders the voting page for the final round.
    :rtype: _TemplateResponse
    """
    voter: Voter = request.state.voter
    rnd = _find_active_round_for_year(db, year, RoundType.FINAL)
    if not rnd:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": f"Активного финала для {year} года нет."},
            status_code=403,
        )
    err = _check_round_open(request, rnd)
    if err:
        return err
    return _render_vote_page(request, db, rnd, voter)


@router.post("/{year}/vote/final")
async def submit_vote_final(year: int, request: Request, db: Session = Depends(get_db)):
    """
    Process submission of votes for a specific year's FINAL round.
    
    Validates the active final round for the given year, checks if voting is open,
    and processes the submitted votes through the _do_submit function.
    
    :param year: The year of the final round for which votes are being submitted.
    :param request: FastAPI request object containing the vote data.
    :param db: The SQLAlchemy database session object.
    :return: Response from _do_submit function processing the vote submission.
    """
    voter: Voter = request.state.voter
    rnd = _find_active_round_for_year(db, year, RoundType.FINAL)
    if not rnd:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": f"Активного финала для {year} года нет."},
            status_code=403,
        )
    err = _check_round_open(request, rnd)
    if err:
        return err
    return await _do_submit(request, db, rnd, voter)

@router.post("/{year}/vote")
async def submit_vote_year(year: int, request: Request, db: Session = Depends(get_db)):
    """
    Compatibility POST endpoint: accept form posts to /{year}/vote and dispatch
    to the currently active round (FINAL or LONGLIST).
    
    :param year: The year for which votes are being submitted.
    :param request: FastAPI request object containing the vote data.
    :param db: The SQLAlchemy database session object.
    :return: Response from _do_submit function processing the vote submission.
    """
    voter: Voter = request.state.voter
    rnd = _find_active_round_for_year(db, year)
    if not rnd:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": f"Активного раунда для {year} года нет."},
            status_code=403,
        )
    err = _check_round_open(request, rnd)
    if err:
        return err
    return await _do_submit(request, db, rnd, voter)


@router.post("/{year}/draft")
async def save_draft_year(year: int, request: Request, db: Session = Depends(get_db)):
    """
    Save a draft for a specific year.
    
    :param year: Year
    :param request: FastAPI request object
    :param db: Database session
    :returns: Dictionary with "ok" and "error" keys
    """
    voter: Voter = request.state.voter
    rnd = _find_active_round_for_year(db, year)
    if not rnd:
        return {"ok": False, "error": "no active round"}
    body = _parse_draft_payload(await request.json())
    p = _get_or_create_participation(db, rnd.id, voter.id)
    p.draft = body
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# POST /{year}/ballot-export
# ---------------------------------------------------------------------------

@router.post("/{year}/ballot-export")
async def ballot_export(year: int, request: Request, db: Session = Depends(get_db)):
    """
    Export a ballot for a specific year.
    
    :param year: Year
    :param request: FastAPI request object
    :param db: Database session
    :returns: StreamingResponse object
    """
    voter: Voter = request.state.voter
    nominations_data = _parse_export_payload(await request.json())
    buf = _build_xlsx_per_nomination(nominations_data)
    safe_label = _safe_filename_label(voter.name or str(voter.id))
    fname = f"ballot_{year}_{safe_label}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(fname)},
    )


# ---------------------------------------------------------------------------
# Compat routes (legacy round_id URLs)
# ---------------------------------------------------------------------------

@router.get("/rounds/{round_id}/vote", response_class=HTMLResponse)
def vote_page(round_id: int, request: Request, db: Session = Depends(get_db)):
    """
    Display the vote page for a specific round.
    
    :param round_id: ID of the round
    :param request: FastAPI request object
    :param db: Database session
    :returns: TemplateResponse object
    """
    rnd = db.get(Round, round_id)
    if not rnd:
        return templates.TemplateResponse(
            request, "voting_closed.html",
            {"nom": None, "message": "Раунд не найден."},
            status_code=404,
        )
    slug = "final" if rnd.round_type == RoundType.FINAL else "longlist"
    return RedirectResponse(url=f"/{rnd.year}/vote/{slug}", status_code=301)


@router.post("/rounds/{round_id}/draft")
async def save_draft(round_id: int, request: Request, db: Session = Depends(get_db)):
    """
    Save a draft for a specific round.
    
    :param round_id: ID of the round
    :param request: FastAPI request object
    :param db: Database session
    :returns: Dictionary with "ok" and "error" keys
    """
    body = _parse_draft_payload(await request.json())
    p = _get_or_create_participation(db, round_id, request.state.voter.id)
    p.draft = body
    db.commit()
    return {"ok": True}


@router.post("/rounds/{round_id}/vote")
async def submit_vote_compat(round_id: int, request: Request, db: Session = Depends(get_db)):
    """
    Submit a vote for a specific round.
    
    :param round_id: ID of the round
    :param request: FastAPI request object
    :param db: Database session
    :returns: Response from _do_submit function processing the vote submission.
    """
    voter: Voter = request.state.voter
    rnd = db.get(Round, round_id)
    err = _check_round_open(request, rnd)
    if err:
        return err
    return await _do_submit(request, db, rnd, voter)


# ---------------------------------------------------------------------------
# Shared submit logic
# ---------------------------------------------------------------------------

async def _do_submit(request: Request, db: Session, rnd: Round, voter: Voter):
    """
    Process the vote submission.
    
    :param request: FastAPI request object
    :param db: Database session
    :param rnd: Round object
    :param voter: Voter object
    :returns: Response from _do_submit function processing the vote submission.
    """
    err = _check_round_open(request, rnd)
    if err:
        return err

    nominations = _nominations_for_round(db, rnd.id)
    form = await request.form()
    is_final = (rnd.round_type.value == "FINAL")

    invalid_runnerups: list[tuple[Nomination, str]] = []
    picks_map: dict[str, list[str]] = {}
    runnerups_map: dict[str, str] = {}
    for nom in nominations:
        if nom.type == NominationType.PICK:
            raw_picks = form.getlist(f"pick_{nom.id}") or []
            picks_set = {str(x) for x in raw_picks}
            picks_map[str(nom.id)] = list(picks_set)
            ru = form.get(f"runnerup_{nom.id}")
            # record runner-up selection (may be empty string when "none" chosen)
            if ru is not None:
                runnerups_map[str(nom.id)] = str(ru)
                if str(ru) != "":
                    if not picks_set:
                        # runner-up chosen but no winner selected
                        invalid_runnerups.append((nom, "no_pick"))
                    elif len(picks_set) == 1 and str(ru) in picks_set:
                        # runner-up equals the sole selected winner
                        invalid_runnerups.append((nom, "same_as_winner"))
                    # otherwise runner-up is acceptable (may be outside picks)
                else:
                    # empty string / none selected — if winner chosen, runner-up is missing
                    if picks_set and nom.has_runner_up:
                        invalid_runnerups.append((nom, "missing_ru"))
            else:
                # field not submitted at all — same as no runner-up chosen
                if picks_set and nom.has_runner_up:
                    invalid_runnerups.append((nom, "missing_ru"))

    ranks_map: dict[str, dict[str, int]] = {}
    errors: list[str] = []

    # Validate picks / runner-ups already collected in picks_map / runnerups_map
    for nom in nominations:
        sid = str(nom.id)
        if nom.type == NominationType.PICK:
            picked = picks_map.get(sid, [])
            # Only validate min/max if the voter actually made a selection (skip = allowed)
            if len(picked) > 0:
                if nom.pick_min and len(picked) < nom.pick_min:
                    errors.append(f"«{nom.name}»: выбрано {len(picked)}, нужно не менее {nom.pick_min}")
                if nom.pick_max and len(picked) > nom.pick_max:
                    errors.append(f"«{nom.name}»: выбрано {len(picked)}, максимум {nom.pick_max}")
            # runner-up already validated above; keep picks_map as-is
        else:
            # collect ranks from form
            rank_entries: list[int] = []
            ranks_map[sid] = {}
            for nominee in nom.nominees:
                val = form.get(f"rank_{nom.id}_{nominee.film_id}")
                if val:
                    try:
                        r = int(val)
                        ranks_map[sid][str(nominee.film_id)] = r
                        rank_entries.append(r)
                    except ValueError:
                        # ignore non-int; will be caught by later checks if necessary
                        pass
            if rank_entries:
                # detect duplicate ranks
                if len(rank_entries) != len(set(rank_entries)):
                    errors.append(f"«{nom.name}»: одно место указано дважды")
                # detect overflow
                max_val = (nom.nominees_count if (not (rnd.round_type.value == "FINAL") and nom.nominees_count) else len(nom.nominees))
                if any(r > max_val for r in rank_entries):
                    errors.append(f"«{nom.name}»: место не может быть больше {max_val}")

    # include runner-up inconsistency errors found earlier
    for nom, code in invalid_runnerups:
        if code == "no_pick":
            errors.append(f"«{nom.name}»: выбран runner-up, но победитель не отмечен")
        elif code == "same_as_winner":
            errors.append(f"«{nom.name}»: победитель и runner-up — один и тот же номинант")
        elif code == "missing_ru":
            errors.append(f"«{nom.name}»: выбран победитель, но runner-up не указан")

    if errors:
        participation = _get_or_create_participation(db, rnd.id, voter.id)
        # include ranks in draft so user sees their inputs
        participation.draft = {"picks": picks_map, "runnerups": runnerups_map, "ranks": ranks_map}
        db.commit()
        return templates.TemplateResponse(request, "vote.html", {
            "voter": voter,
            "round": rnd,
            "nominations": nominations,
            "draft": participation.draft,
            "draft_restored": True,
            "error_message": " · ".join(errors),
        }, status_code=400)

    # Cross-nomination conflict check: prevent voting for same person across acting_group-linked nominations
    selected_nids = set()
    for nom in nominations:
        if nom.type == NominationType.PICK:
            raw = form.getlist(f"pick_{nom.id}") or []
            for v in raw:
                try:
                    selected_nids.add(int(v))
                except ValueError:
                    pass
            if is_final and nom.has_runner_up:
                ru_val = form.get(f"runnerup_{nom.id}")
                if ru_val:
                    try:
                        selected_nids.add(int(ru_val))
                    except ValueError:
                        pass
    try:
        _check_cross_nomination_conflict(db, voter.id, selected_nids, rnd)
    except HTTPException as e:
        participation = _get_or_create_participation(db, rnd.id, voter.id)
        participation.draft = {"picks": picks_map, "runnerups": runnerups_map, "ranks": ranks_map}
        db.commit()
        return templates.TemplateResponse(request, "vote.html", {
            "voter": voter,
            "round": rnd,
            "nominations": nominations,
            "draft": participation.draft,
            "draft_restored": True,
            "error_message": e.detail,
        }, status_code=400)

    round_nominee_ids = {n.id for nom in nominations for n in nom.nominees}

    for nom in nominations:
        if nom.type == NominationType.RANK:
            db.query(Ranking).filter(
                Ranking.voter_id == voter.id,
                Ranking.nomination_id == nom.id,
            ).delete()
        else:
            nominee_ids = [n.id for n in nom.nominees]
            if nominee_ids:
                db.query(Vote).filter(
                    Vote.voter_id == voter.id,
                    Vote.nominee_id.in_(nominee_ids),
                ).delete(synchronize_session="fetch")

    is_final = (rnd.round_type.value == "FINAL")
    # Track nominee_ids added during this submission to avoid duplicate INSERTs
    added_vote_nids: set[int] = set()

    for nom in nominations:
        if nom.type == NominationType.PICK:
            key = f"pick_{nom.id}"
            raw = form.getlist(key)
            if nom.pick_max and len(raw) > nom.pick_max:
                raw = raw[:nom.pick_max]

            # Parse runner-up selection (if any) once per nomination
            ru_nid = None
            if is_final and nom.has_runner_up:
                ru_val = form.get(f"runnerup_{nom.id}")
                if ru_val:
                    try:
                        ru_nid = int(ru_val)
                    except ValueError:
                        ru_nid = None

            raw_nids: set[int] = set()
            for val in raw:
                try:
                    nid = int(val)
                    raw_nids.add(nid)
                    if nid in round_nominee_ids:
                        is_ru = (ru_nid == nid)
                        # Update existing vote if present, otherwise add new (guard against duplicates within same request)
                        existing = db.query(Vote).filter_by(voter_id=voter.id, nominee_id=nid).first()
                        if existing:
                            existing.is_runner_up = existing.is_runner_up or is_ru
                            added_vote_nids.add(nid)
                        else:
                            if nid not in added_vote_nids:
                                db.add(Vote(voter_id=voter.id, nominee_id=nid, is_runner_up=is_ru))
                                added_vote_nids.add(nid)
                except ValueError:
                    pass

            # If runner-up chosen but not among picks, ensure a single runner-up vote exists
            if ru_nid and ru_nid not in raw_nids and ru_nid in round_nominee_ids:
                existing = db.query(Vote).filter_by(voter_id=voter.id, nominee_id=ru_nid).first()
                if existing:
                    existing.is_runner_up = True
                    added_vote_nids.add(ru_nid)
                else:
                    if ru_nid not in added_vote_nids:
                        db.add(Vote(voter_id=voter.id, nominee_id=ru_nid, is_runner_up=True))
                        added_vote_nids.add(ru_nid)
        else:
            for nominee in nom.nominees:
                val = form.get(f"rank_{nom.id}_{nominee.film_id}")
                if val:
                    try:
                        rank = int(val)
                        if 1 <= rank <= len(nom.nominees):
                            db.add(Ranking(
                                voter_id=voter.id,
                                nomination_id=nom.id,
                                film_id=nominee.film_id,
                                rank=rank,
                            ))
                    except ValueError:
                        pass

    participation = _get_or_create_participation(db, rnd.id, voter.id)
    participation.voted_at = datetime.now(timezone.utc)
    participation.draft = None
    db.commit()
    slug = "final" if rnd.round_type == RoundType.FINAL else "longlist"
    return RedirectResponse(url=f"/thank-you?round_id={rnd.id}", status_code=303)


# ---------------------------------------------------------------------------
# /thank-you
# ---------------------------------------------------------------------------

@router.get("/thank-you", response_class=HTMLResponse)
def thank_you(request: Request, db: Session = Depends(get_db)):
    """
    Display the thank you page.
    
    :param request: FastAPI request object
    :param db: Database session
    :returns: TemplateResponse object
    """
    voter: Voter = request.state.voter
    round_id_str = request.query_params.get("round_id")
    has_votes = False
    round_id = None
    if round_id_str:
        try:
            round_id = int(round_id_str)
            p = db.query(RoundParticipation).filter_by(
                round_id=round_id, voter_id=voter.id
            ).first()
            has_votes = bool(p and p.voted_at)
        except ValueError:
            pass
    return templates.TemplateResponse(
        request, "thank_you.html",
        {"has_votes": has_votes, "round_id": round_id},
    )


# ---------------------------------------------------------------------------
# /my-ballot/{round_id}/export
# ---------------------------------------------------------------------------

@router.get("/my-ballot/{round_id}/export")
def export_my_ballot(round_id: int, request: Request, db: Session = Depends(get_db)):
    """
    Export my ballot for a specific round.
    
    :param round_id: ID of the round
    :param request: FastAPI request object
    :param db: Database session
    :returns: StreamingResponse object
    """
    voter: Voter = request.state.voter
    p = db.query(RoundParticipation).filter_by(
        round_id=round_id, voter_id=voter.id
    ).first()
    if not p or not p.voted_at:
        return RedirectResponse(url=f"/thank-you?round_id={round_id}", status_code=303)

    nominations = _nominations_for_round(db, round_id)
    rnd = db.get(Round, round_id)

    nominee_ids = {n.id for nom in nominations for n in nom.nominees}
    pick_votes = {
        v.nominee_id: v.is_runner_up
        for v in db.query(Vote).filter(
            Vote.voter_id == voter.id,
            Vote.nominee_id.in_(nominee_ids),
        ).all()
    }
    rankings = {
        (r.nomination_id, r.film_id): r.rank
        for r in db.query(Ranking).filter(
            Ranking.voter_id == voter.id,
            Ranking.nomination_id.in_({nom.id for nom in nominations}),
        ).all()
    }

    nominations_data = []
    for nom in nominations:
        if nom.type == NominationType.PICK:
            header = ["Номинант", "Голос"]
            rows = []
            for n in nom.nominees:
                if n.id in pick_votes:
                    label = (
                        (n.persons_label and f"{n.persons_label} ({n.film.title})")
                        or (n.item and f"{n.item} ({n.film.title})")
                        or n.film.title
                    )
                    url = (
                        (n.person.url if n.person else None)
                        or getattr(n, 'item_url', None)
                        or (n.film.url if n.film else None)
                    )
                    vote_label = "runner-up" if pick_votes[n.id] else "✔"
                    rows.append({"cols": [label, vote_label], "urls": [url, None]})
            if not rows:
                rows = [{"cols": ["— пропущено", ""], "urls": [None, None]}]
        else:
            header = ["Место", "Фильм"]
            ranked = [
                (
                    rankings[(nom.id, n.film_id)],
                    n.film.title,
                    n.film.url if n.film else None,
                )
                for n in nom.nominees
                if (nom.id, n.film_id) in rankings
            ]
            if ranked:
                rows = [
                    {"cols": [rank, title], "urls": [None, url]}
                    for rank, title, url in sorted(ranked)
                ]
            else:
                rows = [{"cols": ["", "— пропущено"], "urls": [None, None]}]

        nominations_data.append({
            "name": nom.name,
            "type": nom.type.value,
            "header": header,
            "rows": rows,
        })

    buf = _build_xlsx_per_nomination(nominations_data)
    label_safe = _safe_filename_label(rnd.label if rnd else str(round_id))
    voter_safe = _safe_filename_label(voter.name or str(voter.id))
    fname = f"ballot_{voter_safe}_{label_safe}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(fname)},
    )
